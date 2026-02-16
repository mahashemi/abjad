import pandas as pd
import json
import os

cwd = os.getcwd()
debug_dir = os.path.join(cwd, "debug")

# Check if debug directory exists
if not os.path.exists(debug_dir):
    print(f"Error: Debug directory not found at {debug_dir}")
    exit(1)

dirs = os.listdir(debug_dir)
print(f"Found {len(dirs)} directories in debug folder:")
for d in dirs:
    print(f"  - {d}")

# List to store all data including intermediate aya totals
all_data = []

for dir in dirs:
    print(f"\nProcessing directory: {dir}")
    # Handle Arabic directory names like "سورة 98" 
    if dir.startswith('سورة '):
        print(f"  Directory matches Arabic pattern (starts with 'سورة ')")
        # Extract number after "سورة "
        surah_number_str = dir.replace('سورة ', '').strip()
        surah_number = f"{int(surah_number_str):03d}"
        print(f"  Extracted surah number: {surah_number}")
    # Also handle original format with dash (if any exist)
    elif '-' in dir:
        print(f"  Directory matches dash pattern (contains '-')")
        surah_number = f"{int(dir.split('-')[1].strip().split(' ')[1].strip()):03d}"
        print(f"  Extracted surah number: {surah_number}")
    else:
        print(f"  Skipping {dir} (doesn't match expected patterns)")
        continue
    
    # Add separator row for new sura
    separator_row = {
        'type': 'separator',
        'surah_number': surah_number,
        'surah': f"=== {dir} ===",
        'aya_number': '',
        'qamari': '',
        'malfuzi': '',
        'bayenati': ''
    }
    all_data.append(separator_row)
    
    # Initialize totals for this sura
    total_qamari = 0
    total_malfuzi = 0
    total_bayenati = 0
    
    # Initialize subtotals for every 10 ayas
    subtotal_qamari = 0
    subtotal_malfuzi = 0
    subtotal_bayenati = 0
    aya_counter = 0
    
    json_dir = os.path.join(debug_dir, dir)
    if not os.path.exists(json_dir):
        print(f"  Warning: Directory {json_dir} does not exist")
        continue
        
    json_folders = os.listdir(json_dir)
    print(f"  Found {len(json_folders)} items in {dir}:")
    for jf in json_folders:
        print(f"    - {jf}")
    
    # Sort aya folders numerically
    numeric_folders = [f for f in json_folders if f.isnumeric()]
    numeric_folders.sort(key=int)
    print(f"  Numeric folders: {numeric_folders}")
    
    for json_folder in numeric_folders:
        json_file = os.path.join(json_dir, json_folder, 'result.json')
        print(f"    Checking file: {json_file}")
        
        if os.path.exists(json_file):
            print(f"    File exists, reading...")
            try:
                with open(json_file, 'r') as fp:
                    json_data = json.load(fp=fp)
                    print(f"    JSON keys: {list(json_data.keys())}")
                    
                    # Get individual aya values
                    aya_qamari = json_data.get('total_qamari_value', 0)
                    aya_malfuzi = json_data.get('total_malfuzi_value', 0)
                    aya_bayenati = json_data.get('total_bayenati_value', 0)
                    
                    print(f"    Values - Q:{aya_qamari}, M:{aya_malfuzi}, B:{aya_bayenati}")
                    
                    # Update running totals
                    total_qamari += aya_qamari
                    total_malfuzi += aya_malfuzi
                    total_bayenati += aya_bayenati
                    
                    # Update subtotals for every 10 ayas
                    subtotal_qamari += aya_qamari
                    subtotal_malfuzi += aya_malfuzi
                    subtotal_bayenati += aya_bayenati
                    aya_counter += 1
                    
                    # Add aya data row
                    aya_row = {
                        'type': 'aya',
                        'surah_number': surah_number,
                        'surah': dir,
                        'aya_number': int(json_folder),
                        'qamari': aya_qamari,
                        'malfuzi': aya_malfuzi,
                        'bayenati': aya_bayenati
                    }
                    all_data.append(aya_row)
                    
                    # Add subtotal row after every 10 ayas
                    if aya_counter % 10 == 0:
                        subtotal_row = {
                            'type': 'subtotal',
                            'surah_number': surah_number,
                            'surah': f"Subtotal (Ayas {aya_counter - 9}-{aya_counter})",
                            'aya_number': '',
                            'qamari': subtotal_qamari,
                            'malfuzi': subtotal_malfuzi,
                            'bayenati': subtotal_bayenati
                        }
                        all_data.append(subtotal_row)
                        print(f"    Added subtotal for ayas {aya_counter - 9}-{aya_counter} - Q:{subtotal_qamari}, M:{subtotal_malfuzi}, B:{subtotal_bayenati}")
                        
                        # Add break (empty row) after subtotal
                        empty_row = {col: '' for col in subtotal_row.keys()}
                        all_data.append(empty_row)
                        
                        # Reset subtotals
                        subtotal_qamari = 0
                        subtotal_malfuzi = 0
                        subtotal_bayenati = 0
            except Exception as e:
                print(f"    Error reading JSON: {e}")
        else:
            print(f"    File does not exist")
    
    # Add final subtotal if there are remaining ayas (less than 10)
    if aya_counter % 10 != 0 and aya_counter > 0:
        start_aya = (aya_counter // 10) * 10 + 1
        subtotal_row = {
            'type': 'subtotal',
            'surah_number': surah_number,
            'surah': f"Subtotal (Ayas {start_aya}-{aya_counter})",
            'aya_number': '',
            'qamari': subtotal_qamari,
            'malfuzi': subtotal_malfuzi,
            'bayenati': subtotal_bayenati
        }
        all_data.append(subtotal_row)
        print(f"    Added final subtotal for ayas {start_aya}-{aya_counter} - Q:{subtotal_qamari}, M:{subtotal_malfuzi}, B:{subtotal_bayenati}")
        
        # Add break (empty row) after subtotal
        empty_row = {col: '' for col in subtotal_row.keys()}
        all_data.append(empty_row)
        
    # Add sura total row
    total_row = {
        'type': 'total',
        'surah_number': surah_number,
        'surah': f"TOTAL - {dir}",
        'aya_number': '',
        'qamari': total_qamari,
        'malfuzi': total_malfuzi,
        'bayenati': total_bayenati
    }
    all_data.append(total_row)
    print(f"  Added total row - Q:{total_qamari}, M:{total_malfuzi}, B:{total_bayenati}")
    
    # Add empty row for spacing
    empty_row = {col: '' for col in total_row.keys()}
    all_data.append(empty_row)

print(f"\nTotal data rows collected: {len(all_data)}")

# Create DataFrame with all data
df = pd.DataFrame(data=all_data)

# Print available columns for debugging
print("Available columns:", df.columns.tolist())

# Only reorder columns that actually exist
available_columns = df.columns.tolist()
desired_order = ['type', 'surah_number', 'surah', 'aya_number', 'malfuzi', 'bayenati', 'qamari']
column_order = [col for col in desired_order if col in available_columns]

# Add any remaining columns that weren't in our desired order
remaining_cols = [col for col in available_columns if col not in column_order]
column_order.extend(remaining_cols)

df = df[column_order]

# Save to Excel with formatting
with pd.ExcelWriter('verification_detailed.xlsx', engine='openpyxl') as writer:
    df.to_excel(writer, sheet_name='Detailed Analysis', index=False)
    
    # Get the workbook and worksheet
    workbook = writer.book
    worksheet = writer.sheets['Detailed Analysis']
    
    # Apply formatting
    from openpyxl.styles import Font, PatternFill, Alignment
    
    # Header formatting
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    
    for cell in worksheet[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
    
    # Separator row formatting
    separator_font = Font(bold=True, color='000080')
    separator_fill = PatternFill(start_color='E6F3FF', end_color='E6F3FF', fill_type='solid')
    
    # Total row formatting
    total_font = Font(bold=True, color='008000')
    total_fill = PatternFill(start_color='F0FFF0', end_color='F0FFF0', fill_type='solid')
    
    # Subtotal row formatting
    subtotal_font = Font(bold=True, color='FF8C00')
    subtotal_fill = PatternFill(start_color='FFF8DC', end_color='FFF8DC', fill_type='solid')
    
    # Apply formatting to data rows
    # Apply formatting to data rows
    for row_num, row in enumerate(worksheet.iter_rows(min_row=2), start=2):
        if row[0].value == 'separator':
            for cell in row:
                cell.font = separator_font
                cell.fill = separator_fill
                cell.alignment = Alignment(horizontal='center')
        elif row[0].value == 'total':
            for cell in row:
                cell.font = total_font
                cell.fill = total_fill
        elif row[0].value == 'subtotal':
            for cell in row:
                cell.font = subtotal_font
                cell.fill = subtotal_fill
    
    # Adjust column widths
    column_widths = [8, 12, 25, 10, 10, 10, 10]
    for i, width in enumerate(column_widths, 1):
        worksheet.column_dimensions[worksheet.cell(row=1, column=i).column_letter].width = width

print("Detailed verification file created: verification_detailed.xlsx")

# Also create a summary file with just the totals (original format)
summary_data = []
for item in all_data:
    if item.get('type') == 'total':
        summary_data.append({
            'surah_number': item['surah_number'],
            'surah': item['surah'].replace('TOTAL - ', ''),
            'total_qamari': item['qamari'],
            'total_malfuzi': item['malfuzi'],
            'total_bayenati': item['bayenati']
        })

pd.DataFrame(data=summary_data).to_excel('verification_summary.xlsx', index=False)
print("Summary file created: verification_summary.xlsx")